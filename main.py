from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, NamedStyle, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import pandas as pd
import time
import pyautogui
import os
import glob
import random
import xlwings as xw
import pdfkit

# Definir rutas a las carpetas y archivos
input_folder_excel = "C:/Proyectos/-Impuestos/Anticipos-SCT-Tributaria/data/input"

# Leer el archivo Excel
df = pd.read_excel(r'C:/Proyectos/-Impuestos/Anticipos-SCT-Tributaria/data/input/Clientes.xlsx')

# Suposición de nombres de columnas
cuit_login_list = df['CUIT para ingresar'].tolist()
cuit_represent_list = df['CUIT representado'].tolist()
password_list = df['Contraseña'].tolist()
download_list = df['Ubicacion Descarga'].tolist()
posterior_list = df['Posterior'].tolist()
anterior_list = df['Anterior'].tolist()
clientes_list = df['Cliente'].tolist()

# Configuración de opciones de Chrome
options = Options()
options.add_argument("--start-maximized")

# Configurar preferencias de descarga
prefs = {
    "download.prompt_for_download": True,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
options.add_experimental_option("prefs", prefs)

# Inicializar driver
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

# Crear el archivo de resultados
resultados = []

def human_typing(element, text):
    for char in str(text):
        element.send_keys(char)
        time.sleep(random.uniform(0.05, 0.3))

def actualizar_excel(row_index, mensaje):
    """Actualiza la última columna del archivo Excel con un mensaje de error."""
    df.at[row_index, 'Error'] = mensaje
    df.to_excel(r'C:/Proyectos/-Impuestos/Anticipos-SCT-Tributaria/data/input/Clientes.xlsx', index=False)

def iniciar_sesion(cuit_ingresar, password, row_index):
    """Inicia sesión en el sitio web con el CUIT y contraseña proporcionados."""
    try:
        driver.get('https://auth.afip.gob.ar/contribuyente_/login.xhtml')
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:username')))
        element.clear()
        time.sleep(5)

        human_typing(element, cuit_ingresar)
        driver.find_element(By.ID, 'F1:btnSiguiente').click()
        time.sleep(5)

        # Verificar si el CUIT es incorrecto
        try:
            error_message = driver.find_element(By.ID, 'F1:msg').text
            if error_message == "Número de CUIL/CUIT incorrecto":
                actualizar_excel(row_index, "Número de CUIL/CUIT incorrecto")
                return False
        except:
            pass

        element_pass = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:password')))
        human_typing(element_pass, password)
        time.sleep(15)
        driver.find_element(By.ID, 'F1:btnIngresar').click()
        time.sleep(5)

        # Verificar si la contraseña es incorrecta
        try:
            error_message = driver.find_element(By.ID, 'F1:msg').text
            if error_message == "Clave o usuario incorrecto":
                actualizar_excel(row_index, "Clave o usuario incorrecto")
                return False
        except:
            pass

        return True
    except Exception as e:
        print(f"Error al iniciar sesión: {e}")
        actualizar_excel(row_index, "Error al iniciar sesión")
        return False

def ingresar_modulo(cuit_ingresar, password, row_index):
    """Ingresa al módulo específico del sistema de cuentas tributarias."""
    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, "Ver todos"))).click()
        time.sleep(5)

        element = driver.find_element(By.ID, 'buscadorInput')
        human_typing(element, 'tas tr') 
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'rbt-menu-item-0'))).click()
        time.sleep(10)

        try:
            # Esperar y manejar el modal si aparece
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, 'modal-content')))
            modal = driver.find_element(By.CLASS_NAME, 'modal-content')
            if modal.is_displayed():
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Continuar"]'))).click()
                time.sleep(5)
        except:
            # No hacer nada si el modal no aparece
            pass

        # Cambiar de pestaña
        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[-1])

        # Verificar mensaje de error de autenticación
        try:
            error_message = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, 'pre')))
            if error_message.text == "Ha ocurrido un error al autenticar, intente nuevamente.":
                actualizar_excel(row_index, "Error autenticacion")
                driver.refresh()
                time.sleep(5)
        except:
            pass

        # Verificar si es necesario iniciar sesion nuevamente
        try:
            element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:username')))
            element.clear()
            time.sleep(5)

            human_typing(element, cuit_ingresar)
            driver.find_element(By.ID, 'F1:btnSiguiente').click()
            time.sleep(5)

            element_pass = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:password')))
            human_typing(element_pass, password)
            time.sleep(15)
            driver.find_element(By.ID, 'F1:btnIngresar').click()
            time.sleep(5)

            actualizar_excel(row_index, "Error volver a iniciar sesion")
        except:
            pass

    except Exception as e:
        print(f"Error al ingresar al módulo: {e}")

def seleccionar_cuit_representado(cuit_representado):
    """Selecciona el CUIT representado en el sistema."""
    try:
        select_present = EC.presence_of_element_located((By.NAME, "$PropertySelection"))
        if WebDriverWait(driver, 5).until(select_present):
            current_selection = Select(driver.find_element(By.NAME, "$PropertySelection")).first_selected_option.text
            if current_selection != str(cuit_representado):
                select_element = Select(driver.find_element(By.NAME, "$PropertySelection"))
                select_element.select_by_visible_text(str(cuit_representado))
    except Exception:
        try:
            cuit_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'span.cuit')))
            cuit_text = cuit_element.text.replace('-', '')
            if cuit_text != str(cuit_representado):
                print(f"El CUIT ingresado no coincide con el CUIT representado: {cuit_representado}")
                return False
        except Exception as e:
            print(f"Error al verificar CUIT: {e}")
            return False
    # Esperar que el popup esté visible y hacer clic en el botón de cerrar por XPATH
    try:
    # Usamos el XPATH para localizar el botón de cerrar
        close_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, '//a[@href="#close" and @title="Cerrar"]'))
        )
        close_button.click()
        print("Popup cerrado exitosamente.")
    except Exception as e:
        print(f"Error al intentar cerrar el popup: {e}")
    return True

def exportar_excel(ubicacion_descarga, cuit_representado, cliente, cantidad_faltas_presentacion):
    """Descarga y guarda el archivo Excel en la ubicación especificada."""
    try:       
        # Exportar XLSX
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='DataTables_Table_0_wrapper']/div[1]/a[2]/span"))).click()
        time.sleep(5)

        # Guardarlo con nombre y carpeta especifica

        nombre_archivo = f"Anticipos - {cliente}"
        pyautogui.write(nombre_archivo)
        time.sleep(1)
        pyautogui.hotkey('alt', 'd')
        time.sleep(0.5)
        pyautogui.write(ubicacion_descarga)
        time.sleep(1)
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.hotkey('alt', 't')
        time.sleep(1)
        pyautogui.press('enter')
        time.sleep(1)
    except Exception as e:
        print(f"Error al exportar el archivo Excel: {e}")

def cerrar_sesion():
    """Cierra la sesión actual."""
    try:
        driver.close()
        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[0])
        driver.find_element(By.ID, "iconoChicoContribuyenteAFIP").click()
        driver.find_element(By.XPATH, '//*[@id="contBtnContribuyente"]/div[6]/button/div/div[2]').click()
        time.sleep(5)
    except Exception as e:
        print(f"Error al cerrar sesión: {e}")

def extraer_datos_nuevo(cuit_ingresar, cuit_representado, password, ubicacion_descarga, posterior, cliente, indice):
    """Extrae datos para un nuevo usuario."""
    try:
        control_sesion = iniciar_sesion(cuit_ingresar, password, indice)
        if control_sesion:
            ingresar_modulo(cuit_ingresar, password, indice)
            # Esperar que el popup esté visible y hacer clic en el botón de cerrar por XPATH
            try:
                # Usamos el XPATH para localizar el botón de cerrar
                close_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, '//a[@href="#close" and @title="Cerrar"]'))
                )
                close_button.click()
                print("Popup cerrado exitosamente.")
            except Exception as e:
                print(f"Error al intentar cerrar el popup: {e}")
            if seleccionar_cuit_representado(cuit_representado):
                cantidad_faltas_presentacion = driver.find_element(By.NAME, "functor$1").get_attribute('value')
                exportar_excel(ubicacion_descarga, cuit_representado, cliente, cantidad_faltas_presentacion)
                if posterior == 0:
                    cerrar_sesion()
                return cantidad_faltas_presentacion
    except Exception as e:
        print(f"Error al extraer datos para el nuevo usuario: {e}")

def extraer_datos(cuit_representado, ubicacion_descarga, posterior, cliente):
    """Extrae datos para un usuario existente."""
    try:
        if seleccionar_cuit_representado(cuit_representado):
            cantidad_faltas_presentacion = driver.find_element(By.NAME, "functor$1").get_attribute('value')
            exportar_excel(ubicacion_descarga, cuit_representado, cliente, cantidad_faltas_presentacion)
            if posterior == 0:
                cerrar_sesion()
    except Exception as e:
        print(f"Error al extraer datos: {e}")

# Función para convertir Excel a CSV utilizando xlwings
def excel_a_csv(input_folder, output_folder):
    for excel_file in glob.glob(os.path.join(input_folder, "*.xlsx")):
        try:
            app = xw.App(visible=False)
            wb = app.books.open(excel_file)
            sheet = wb.sheets[0]
            df = sheet.used_range.options(pd.DataFrame, header=1, index=False).value

            # Convertir la columna 'FechaVencimiento' a datetime, ajustar según sea necesario
            if 'FechaVencimiento' in df.columns:
                df['FechaVencimiento'] = pd.to_datetime(df['FechaVencimiento'], errors='coerce')

            wb.close()
            app.quit()

            base = os.path.basename(excel_file)
            csv_file = os.path.join(output_folder, base.replace('.xlsx', '.csv'))
            df.to_csv(csv_file, index=False, encoding='utf-8-sig', sep=';')
            print(f"Convertido {excel_file} a {csv_file}")
        except Exception as e:
            print(f"Error al convertir {excel_file} a CSV: {e}")

# Función para obtener el nombre del cliente a partir del nombre del archivo
def obtener_nombre_cliente(filename):
    base = os.path.basename(filename)
    nombre_cliente = base.split('-')[1].strip()
    return nombre_cliente

# Función para obtener la cantidad de faltas de presentación a partir del nombre del archivo
def obtener_faltas_presentacion(filename):
    base = os.path.basename(filename)
    faltas_presentacion = int(base.split('-')[2].strip())
    return faltas_presentacion

# Iterar sobre cada cliente
indice = 0
for cuit_ingresar, cuit_representado, password, download, posterior, anterior, cliente in zip(cuit_login_list, cuit_represent_list, password_list, download_list, posterior_list, anterior_list, clientes_list):
    if anterior == 0:
        extraer_datos_nuevo(cuit_ingresar, cuit_representado, password, download, posterior, cliente, indice)
    else:
        extraer_datos(cuit_representado, download, posterior, cliente)
    indice = indice + 1

output_folder_pdf = "C:/Proyectos/-Impuestos/ANTICIPOS-SCT-TRIBUTARIA/data/Reportes"
imagen = "C:/Proyectos/-Impuestos/ANTICIPOS-SCT-TRIBUTARIA/data/imagen.png"

def forzar_guardado_excel(excel_file):
    app = xw.App(visible=False)
    wb = app.books.open(excel_file)
    wb.save()
    wb.close()
    app.quit()

def procesar_excel(excel_file, output_pdf, imagen):
    # Cargar el archivo Excel con pandas
    df = pd.read_excel(excel_file)

    # Filtrar por "Periodo fiscal" y "Impuesto"
    df_filtrado = df[
        (df['Período Fiscal'].astype(str).str.contains('2025')) & 
        (df['Impuesto'].str.contains('ganancias personas físicas|ganancias personas fisicas|bienes personales', case=False, na=False))
    ]

    # Verificar si la tabla está vacía
    if df_filtrado.shape[0] == 0:
        output_pdf = output_pdf.replace(".pdf", " - vacio.pdf")

    # Eliminar las columnas innecesarias
    if 'Concepto / Subconcepto' in df.columns:
        df_filtrado = df_filtrado.drop(['Concepto / Subconcepto'], axis=1)

    if 'Int. resarcitorios' in df.columns:
        df_filtrado = df_filtrado.drop(['Int. resarcitorios'], axis=1)

    if 'Int. punitorios' in df.columns:
        df_filtrado = df_filtrado.drop(['Int. punitorios'], axis=1)

    # Guardar el DataFrame filtrado en el archivo Excel
    df_filtrado.to_excel(excel_file, index=False)
    
    # Cargar el archivo para aplicar formato con openpyxl
    wb = load_workbook(excel_file)
    ws = wb.active

    # Insertar filas adicionales si necesitas espacio adicional para una nueva imagen
    ws.insert_rows(1, amount=7)

    # Agregar una imagen encima del encabezado (A1)
    img = ExcelImage(imagen)
    total_width = sum([ws.column_dimensions[get_column_letter(col)].width for col in range(1, ws.max_column + 1)])
    img.width = total_width * 7
    img.height = 120
    ws.add_image(img, 'A1')

    # Cambiar el color del encabezado a lila
    header_fill = PatternFill(start_color="AA0EAA", end_color="AA0EAA", fill_type="solid")
    for cell in ws[8]:
        cell.fill = header_fill

    # Ajustar el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Centrar el contenido de todas las celdas
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Guardar los cambios
    wb.save(excel_file)

    # Iniciar una instancia de xlwings de forma invisible
    app = xw.App(visible=False)
    try:
        wb_xw = app.books.open(excel_file)
        last_row = ws.max_row
        last_col = ws.max_column
        wb_xw.sheets[0].api.PageSetup.PrintArea = f"A1:{get_column_letter(last_col)}{last_row}"

        wb_xw.sheets[0].api.PageSetup.Zoom = False
        wb_xw.sheets[0].to_pdf(output_pdf)
        wb_xw.close()
    finally:
        app.quit()

    print(f"Archivo {excel_file} procesado y guardado como {output_pdf}")


# Recorrer todos los archivos Excel en la carpeta
for excel_file in glob.glob(os.path.join(input_folder_excel, "*.xlsx")):
    try:
    # Forzar guardado para evitar problemas con archivos corruptos o no calculados
        forzar_guardado_excel(excel_file)

        # Obtener el nombre base del archivo para usarlo en el nombre del PDF
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_pdf = os.path.join(output_folder_pdf, f"{base_name}.pdf")
        
        # Llamar a la función para procesar el archivo Excel y generar el PDF
        procesar_excel(excel_file, output_pdf, imagen)
        
        print(f"Archivo {excel_file} procesado y guardado como {output_pdf}")
    
    except Exception as e:
        print(f"Error al procesar {excel_file}: {e}")